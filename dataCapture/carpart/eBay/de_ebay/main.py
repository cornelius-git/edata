from openpyxl import load_workbook
import basedata
from multiprocessing import  Process,freeze_support

base_url = "{}/sch/i.html?_from=R40&_nkw={}&_sacat=0&LH_TitleDesc=0&_ipg=200&_fcid={}&_pgn={}"
def lmain(start):
    excel_name = load_workbook("carpartsoe.xlsx")
    sheet_name_list = excel_name.sheetnames
    sheet_active = excel_name["Sheet1"]
    total_row = sheet_active.max_row
    total_columns = sheet_active.max_column
    print(total_row)
    for i in range(start,start+300+1):
        result = sheet_active.cell(row=i,column=1).value
        if result:
            basedata.spiderData(url=base_url,inner=result)
        else:
            break

# if __name__ == "__main__":
#     lmain(start=2)

if __name__ == '__main__':
    freeze_support()
    p1 = Process(target=lmain,args=(3600,))
    p2 = Process(target=lmain, args=(3300,))
    p3 = Process(target=lmain,args=(3000,))
    p1.start()
    p2.start()
    p3.start()



