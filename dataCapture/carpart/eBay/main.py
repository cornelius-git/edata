from openpyxl import load_workbook
import basedata
from helpInput import connectmysql as cn

base_url = "{}/sch/i.html?_from=R40&_nkw={}&_sacat=0&LH_TitleDesc=0&_ipg=200&_fcid={}&_pgn={}"
def lmain(start):
    excel_name = load_workbook("carpartsoe.xlsx")
    sheet_name_list = excel_name.sheetnames
    sheet_active = excel_name["Sheet1"]
    total_row = sheet_active.max_row
    total_columns = sheet_active.max_column
    print(total_row)
    for i in range(start,total_row):
        result = sheet_active.cell(row=i,column=1).value
        res = cn.dataLook(oe=result)
        if res:
            continue
        else:
            basedata.spiderData(url=base_url,inner=result)

if __name__ == "__main__":
    lmain(start=2134)