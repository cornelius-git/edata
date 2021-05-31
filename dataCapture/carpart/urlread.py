from openpyxl import load_workbook
import carpartdata as cp
import pymysql  # 导入 pymysql

# 打开数据库连接
db = pymysql.connect(host="127.0.0.1", user="root",
                     password="root", db="control arm", port=3306,charset='utf8mb4')

# 使用cursor()方法获取操作游标
cur = db.cursor()

class dataoprate:
    def __init__(self):
        self.sum = None
        self.oe = None
    def excelRead(self,start):
        wb = load_workbook("carpart.xlsx")
        ws = wb['Sheet1']
        row_max = ws.max_row
        for i in range(start,row_max):
            title = ws.cell(row=i,column=2).value
            url = ws.cell(row=i,column=3).value
            rev = ws.cell(row=i,column=4).value
            print(title,url,rev)
            self.sum,self.oe = cp.spiderData(url=url)
            print(url)
            self.sqlread(title=title,rev=rev,url=url)


    def textWrite(self,title,url,rev):
        with open('carpart.text',encoding='utf-8',mode='a+') as f:
            inner_text = str(title)+'@'+str(url)+'@'+str(rev)+'@'+str(self.sum)+'@'+str(self.oe)+'\n'
            f.write(inner_text)

    def sqlread(self,title,url,rev):
        sql = "insert into carpart(title,url,rev,sum_inner,oem) values(%s,%s,%s,%s,%s)"
        try:
            res = cur.execute(sql, (title,url,rev,self.sum,self.oe))  # 执行sql语句
            db.commit()
        except Exception as e:
            print(e)
        finally:
            pass

if __name__ == '__main__':
    dataoprate().excelRead(start=2)