
import pymysql  # 导入 pymysql

# 打开数据库连接
db = pymysql.connect(host="127.0.0.1", user="root",
                     password="root", db="controlarm", port=3306,charset='utf8mb4')

# 使用cursor()方法获取操作游标
cur = db.cursor()
import logging
import time

from selenium.webdriver import ChromeOptions
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

#输出日志的
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s: %(message)s')
TIME_OUT = 30
TOTAL_PAGE = 12
#操作浏览器的基础配置
option = ChromeOptions()
#隐藏自动化程序
option.add_experimental_option('excludeSwitches', ['enable-automation'])
option.add_experimental_option('useAutomationExtension', False)
#限制图片加载
prefs={
    'profile.default_content_setting_values': {
        'images': 2}
}
option.add_experimental_option('prefs',prefs)

#开启无界面模式
# option.add_argument("--headless")
# option.add_argument("--disable-gpu")
# wait = WebDriverWait(browser, TIME_OUT)



# 'https://www.ebay.de':77,
# market_dict = {'https://www.ebay.com':1,'https://www.ebay.co.uk':3,'https://www.ebay.com.au':15,}
market_dict = {'https://www.ebay.de':77}
# 1.查询操作
# 编写sql 查询语句  user 对应我的表名
def dataSql(itemid,title,picture,oe,market,judege,sold):
    sql = "insert into baseinfo_copy(itemid,title,picture,oe,market,judege,sold) values(%s,%s,%s,%s,%s,%s,%s)"
    try:
        res = cur.execute(sql,(itemid,title,picture,oe,market,judege,sold))  # 执行sql语句
        db.commit()
    except Exception as e:
        print(e)
    finally:
        pass
        # cur.close()  # 关闭连接

class helpIn:
    def __init__(self,searchid=1):
        self.searchkey = searchid
        self.url = '{}/sch/i.html?_from=R40&_nkw={}&_sacat=0&LH_TitleDesc=0&_ipg=200&_fcid={}&_pgn={}&LH_ItemCondition=1000'
        self.name= None
        self.searchword = None
    def dataRead(self,start,end):
        wb = load_workbook('TURBOde.xlsx')
        wt = wb["Sheet1"]
        for sdf in range(start,end):
            self.searchword = wt.cell(row=sdf,column=1).value
            self.spiderData()

    def dataWrite(self,itemid,title,picture,price,sold):
        #写入数据库
        sql = 'insert into turbode(itemid,title,picture,name,searchword,price,sold) values(%s,%s,%s,%s,%s,%s,%s)'
        cur.execute(sql,(itemid,title,picture,self.name,self.searchword,price,sold))
        db.commit()
    def spiderData(self):
        ert = open("turbode.text",mode="a+",encoding='utf-8')
        print("数据自动化程序即将执行，请不要关闭自动打开的浏览器")
        browser = webdriver.Chrome(options=option, executable_path='../chromedriver.exe')
        # 爬取数据
        browser.delete_all_cookies()
        search_name = str(self.searchword).replace(" ", "+")
        market = []
        # 判断市场
        for mark_url, mark_id in market_dict.items():
            area = mark_url.split('.')[-1]
            if area == "com":
                area = "us"
            else:
                pass
            self.name =area

            for page in range(0, 10):
                # 页面内容提取
                print(page)
                browser.get(self.url.format(mark_url, search_name, mark_id, page + 1))
                time.sleep(1)
                i = 1
                try:
                    count = browser.find_element_by_xpath(
                        '//*[@id="mainContent"]/div[1]/div/div[3]/div[1]/div[1]/h1/span[1]').text
                    count = int(count)
                    jcount = count
                except Exception as yut:
                    break
                ert.write(str(self.searchword)+"#" + str(count)+"\n")
                # 计数结果判断
                if count == 0:
                    break
                elif count > 200 and page > 0:
                    count = count - 200*page

                for i in range (1,count+1):
                    # 循环提取每页内容
                    price = None
                    look_count_result = "no data"
                    try:
                        commodity_name = browser.find_element_by_xpath(
                            '//*[@id="srp-river-results"]/ul/li[{}]/div/div[2]/a/h3'.format(i)).text
                        commodity_link = browser.find_element_by_xpath(
                            '//*[@id="srp-river-results"]/ul/li[{}]/div/div[2]/a'.format(i)).get_attribute("href")
                        image_link = browser.find_element_by_xpath(
                            '//*[@id="srp-river-results"]/ul/li[{}]/div/div[1]/div/a/div/img'.format(i)).get_attribute(
                            "src")
                        price = browser.find_element_by_xpath('//*[@id="srp-river-results"]/ul/li[{}]//span[@class="s-item__price"]'.format(i)).text
                        try:
                            # 尝试抓取销售数据
                            look_count = browser.find_elements_by_xpath(
                                '//*[@id="srp-river-results"]/ul/li[{}]/div/div[2]//span[contains(@class,"BOLD")]'.format(
                                    i))
                            if look_count:
                                look_count_result = ""
                                for gh in look_count:
                                    look_count_result += gh.text
                        except Exception as e:
                            pass
                        # print(i)
                        i += 1
                        # print(name,look_count,look_url)
                        judege = "no"
                        itemid = commodity_link.split("/")[-1].split('?')[0]
                        self.dataWrite(itemid=itemid,title=commodity_name,picture=image_link,price=price,sold=look_count_result)
                    except Exception as y:
                        # browser.find_element_by_xpath('//*[@id="gh-ac"]').send_keys(inner)
                        # browser.find_element_by_xpath('//*[@id="gh-btn"]').click()
                        print(y)
                        print("第一页内容提取完毕")
                        break
                if count < 200:
                    break
                try:
                    # 页面class值是唯一值
                    next_page = browser.find_element_by_xpath('//*[contains(@class,"pagination__next")]')
                    # print(next_page)
                    next_page_judge = next_page.get_attribute('aria-disabled')
                    print(next_page_judge)
                    if next_page_judge:
                        break
                    else:
                        continue
                except Exception as q:
                    print(str(self.searchword) + '采集完毕' + '\t' + '无下一页')
                    break
            # print("执行完毕")
        browser.quit()
        ert.close()
    def run(self,start,end):
       self.dataRead(start=start,end=end)






if __name__ == '__main__':
    helpIn(searchid=1).run(start=2075,end=2076)