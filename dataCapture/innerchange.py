from openpyxl import load_workbook,Workbook
import random

class excelSpilt():
    def __init__(self,file=None):
        self.file = None
        self.oexcel = Workbook()
        self.wexcel = self.oexcel.active
    def rowChange(self,file,start=1):
        ws = load_workbook(file)
        wb = ws['Sheet1']
        wb_max_row = wb.max_row
        print(wb_max_row)
        for i in range(start,wb_max_row+1):
            first_value = wb.cell(row=i,column=1).value
            second_value = wb.cell(row=i,column=2).value
            print(first_value,second_value)
            if ";" in str(second_value):
                value_spilt = second_value.split(";")
                for inner in value_spilt:
                    if str(inner).strip(" "):
                        # print(inner)
                        self.exceleWrite(first_value=first_value,inner=inner)
            else:
                self.exceleWrite(first_value=first_value,inner=second_value)
        self.oexcel.save("{}split.xlsx".format(file.split(".")[0]))
    def exceleWrite(self,first_value,inner):
        emax_row = self.wexcel.max_row
        self.wexcel.cell(row=emax_row+1,column=1,value=first_value)
        self.wexcel.cell(row=emax_row+1,column=2,value=inner)



class excleMerg():
    def __init__(self):
        self.olist = []
        self.first_name = None
        self.second_name =None
        self.oexcel = Workbook()
        self.wexcel = self.oexcel.active

    def  rowChange(self,file,start=1):
        wb = load_workbook(file)
        ws = wb['Sheet1']
        ws_max_row = ws.max_row
        print(ws_max_row)
        for gh in range(start,ws_max_row+1):
            first_value = ws.cell(row=gh,column=1).value
            second_value = ws.cell(row=gh,column=2).value
            if first_value in self.olist:
                continue
            else:
                self.olist.append(first_value)
                for yu in range(gh+1,ws_max_row+1):
                    if first_value == str(ws.cell(row=yu,column=1).value).strip(" "):
                        second_value = str(second_value) + ";" + str(ws.cell(row=yu,column=2).value).strip(" ")
                    else:
                        continue
                self.exceleWrite(first_value=first_value,inner=second_value)
        self.oexcel.save('merge/{}mer.xlsx'.format(file.split(".")[0]))

    def exceleWrite(self, first_value, inner):
        emax_row = self.wexcel.max_row
        self.wexcel.cell(row=emax_row + 1, column=1, value=first_value)
        self.wexcel.cell(row=emax_row+1, column=2, value=inner)

if __name__ == '__main__':
    excelSpilt().rowChange(file='carpartsdata.xlsx')
    # excleMerg().rowChange(file='maxoechange.xlsx')


