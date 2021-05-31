import logging
import time
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver import ChromeOptions
from selenium import webdriver
from openpyxl import load_workbook
# import Excleoperate as eo


#操作浏览器的基础配置
option = ChromeOptions()
#隐藏自动化程序
option.add_experimental_option('excludeSwitches', ['enable-automation'])
option.add_experimental_option('useAutomationExtension', False)
#限制图片加载
prefs={
    'profile.default_content_setting_values': {
        'images': 2
        ,'javascript':2}
}
option.add_experimental_option('prefs',prefs)

#开启无界面模式
# option.add_argument("--headless")
# option.add_argument("--disable-gpu")
# wait = WebDriverWait(browser, TIME_OUT)

# market_dict = {'https://www.ebay.com':1,'https://www.ebay.co.uk':3,'https://www.ebay.com.au':15}
#区分市场
market_dict = {'https://www.ebay.de':77}

#主程序

def readExcel(start):
    ws = load_workbook("5月开发-3A.xlsx")
    wb = ws["Sheet1"]
    row_max =wb.max_row
    for hj in range(start,row_max):
        fg = wb.cell(row=hj,column=1).value
        spiderData(base_url,str(fg))



def spiderData(url,inner):
    print("数据自动化程序即将执行，请不要关闭自动打开的浏览器")
    browser = webdriver.Chrome(options=option, executable_path='../chromedriver.exe')
    #爬取数据
    search_name = str(inner).replace(" ","+")
    market = []
    for mark_url,mark_id in market_dict.items():
        browser.get(url.format(mark_url, search_name, mark_id,))
        area = mark_url.split('.')[-1]
        if area == "com":
            area = "us"
        else:
            pass
        result_ms = 'no data'
        time.sleep(1)

        try :
            # print(name,look_count,look_url)
            result_ms = browser.find_element_by_xpath('//*[@id="mainContent"]/div[1]/div/div[2]/div[1]/div[1]/h1').text

            write_text = str(inner)+"@"+str(area)+"@"+str(result_ms)+"\n"
            # print("第{}页内容".format(page))
            print(write_text)
            with open("Kannade.text".format(inner),mode="a+",encoding='utf-8') as f:
                f.write(write_text)
                f.close()
        except Exception as y:
            # print(y)
            try:
                # print(name,look_count,look_url)
                result_ms = browser.find_element_by_xpath(
                    '//*[@id="mainContent"]/div[1]/div/div[3]/div[1]/div[1]/h1').text

                write_text = str(inner) + "@" + str(area) + "@" + str(result_ms) + "\n"
                # print("第{}页内容".format(page))
                print(write_text)
                with open("Kannade.text".format(inner), mode="a+", encoding='utf-8') as f:
                    f.write(write_text)
                    f.close()
            except Exception as tqw:
                print(tqw)

            print("第一页内容提取完毕")
            # break
    browser.quit()



if __name__ == "__main__":
    base_url = "{}/sch/i.html?_from=R40&_nkw={}&_sacat=0&LH_TitleDesc=0&_ipg=200&_fcid={}"
    readExcel(start=1)
    # spiderData(url=base_url,inner="W02-358-7001")