from selenium.webdriver import ChromeOptions
from selenium import webdriver
from datetime import datetime
from openpyxl import load_workbook
import connectsql as cn
#操作浏览器的基础配置
option = ChromeOptions()
#隐藏自动化程序
option.add_experimental_option('excludeSwitches', ['enable-automation'])
option.add_experimental_option('useAutomationExtension', False)
#限制图片加载
prefs={
    'profile.default_content_setting_values': {
        'images': 2
        # ,"javascript":2
    }
}
option.add_experimental_option('prefs',prefs)



#开启无界面模式
# option.add_argument("--headless")
# option.add_argument("--disable-gpu")
# wait = WebDriverWait(browser, TIME_OUT)
now_time =datetime.now()

# use_agent = option.add_argument('user-agent=' + "self.ua")
# use_agent_list = ["dfg", "uio", "tuty", "yuyy", "sdfsd", "dfsd", "sdfs", "sds", "sdas", "sds"]
# option.add_argument('user-agent=' + str(use_agent_list[random.randint(0, 4)]))
browser = webdriver.Chrome(options=option, executable_path='../chromedriver.exe')


class baseinfo:
    def __init__(self):
        self.itemid = None
        self.inumber = None
        self.oem = None
        self.other_number = None
        self.price = None
        self.sold_number = None
        self.seller = None
        self.seller_url = None
    def urlRequest(self):
        url = 'https://www.ebay.com/itm/{}'.format(self.itemid)
        browser.get(url=url)
        browser.delete_all_cookies()
        descripation = None
        try:
            self.price = browser.find_element_by_xpath('//*[@id="prcIsum"]').text
            self.sold_number = browser.find_element_by_xpath('//*[@id="mainContent"]//span[contains(@class,"vi-qty-pur-lnk")]').text
            # self.static = browser.find_element_by_xpath('//*[@id="why2buy"]/div/div[1]/span').text
            self.seller = browser.find_element_by_xpath('//*[@id="RightSummaryPanel"]//div[@class="bdg-90"]/div[1]/a').text
            self.seller_url = browser.find_element_by_xpath('//*[@id="RightSummaryPanel"]//div[@class="bdg-90"]/div[1]/a').get_attribute('href')
            # print(self.price,self.sold_number,self.seller,self.seller_url)
            descripation = browser.find_element_by_xpath('//*[@id="viTabs_0_is"]').text
            # car_sum = browser.find_element_by_xpath('//*[@id="w1-30pgn"]/div/span[1]').text
        except Exception as t:
            pass
        i = 1
        while True:
            try:
                tr  = browser.find_element_by_xpath('//*[@id="viTabs_0_is"]/div/table/tbody/tr[{}]'.format(i)).text
                for f in range(1,5):
                    judege = str(browser.find_element_by_xpath('//*[@id="viTabs_0_is"]/div/table/tbody/tr[{}]/td[{}]'.format(i,f)).text).upper()
                    try:
                        if f == 1 or f == 3:
                            if 'NUMBER' in judege:
                                if "INTER" in judege:
                                    self.inumber = browser.find_element_by_xpath('//*[@id="viTabs_0_is"]/div/table/tbody/tr[{}]/td[{}]'.format(i,f+1)).text
                                elif "MANU" in judege:
                                    self.oem = browser.find_element_by_xpath('//*[@id="viTabs_0_is"]/div/table/tbody/tr[{}]/td[{}]'.format(i,f+1)).text
                                else:
                                    self.other_number = browser.find_element_by_xpath('//*[@id="viTabs_0_is"]/div/table/tbody/tr[{}]/td[{}]'.format(i,f+1)).text
                            else:
                                continue
                        else:
                            continue
                    except Exception as e:
                        print("错误1",e)
                        break
                i += 1
            except Exception as y:
                print("错误2",y)
                break
        if self.price:
            cn.saleInfoSpecific(itemid=self.itemid,store_link=self.seller_url,current_price=self.price,
                            description=descripation,store_name=self.seller,oem=self.oem,internumber=self.inumber,othernumber=self.other_number,sold=self.sold_number)

        # print(self.inumber,self.oem,self.other_number)

    def fileRead(self,start,end):
        ws = load_workbook('datasummary.xlsx')
        wt = ws['Sheet1']
        for wr in range(start,end):
            self.itemid = wt.cell(row=wr,column=2).value
            self.inumber = None
            self.oem = None
            self.other_number = None
            self.price = None
            self.sold_number = None
            self.seller = None
            self.seller_url = None
            print(self.itemid)
            self.urlRequest()






if __name__ == '__main__':
    baseinfo().fileRead(start=6561,end=8322)




