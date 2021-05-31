import requests,datetime,re
from openpyxl import load_workbook
from lxml import etree
import random
time_dict = {'JAN':'01','FEB':'02','MAR':'03','APR':'04','MAY':'05','JUN':'06',
             'JUL':'07','AUG':'08','SEP':'09','OCT':'10','NOV':'11','DEC':'12'}
# import socket
# import socks
import connectsql as cn
proxy = '127.0.0.1:10808'
proxies = {
    'http': 'socks5://' + proxy,
    'https': 'socks5://' + proxy
}


user_agent = [
    "Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_8; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50",
    "Mozilla/5.0 (Windows; U; Windows NT 6.1; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50",
    "Mozilla/5.0 (Windows NT 10.0; WOW64; rv:38.0) Gecko/20100101 Firefox/38.0",
    "Mozilla/5.0 (Windows NT 10.0; WOW64; Trident/7.0; .NET4.0C; .NET4.0E; .NET CLR 2.0.50727; .NET CLR 3.0.30729; .NET CLR 3.5.30729; InfoPath.3; rv:11.0) like Gecko",
    "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0)",
    "Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0)",
    "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1)",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.6; rv:2.0.1) Gecko/20100101 Firefox/4.0.1",
    "Mozilla/5.0 (Windows NT 6.1; rv:2.0.1) Gecko/20100101 Firefox/4.0.1",
    "Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; en) Presto/2.8.131 Version/11.11",
    "Opera/9.80 (Windows NT 6.1; U; en) Presto/2.8.131 Version/11.11",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_0) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Maxthon 2.0)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; TencentTraveler 4.0)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; The World)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; SE 2.X MetaSr 1.0; SE 2.X MetaSr 1.0; .NET CLR 2.0.50727; SE 2.X MetaSr 1.0)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; 360SE)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Avant Browser)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1)",
    "Mozilla/5.0 (iPhone; U; CPU iPhone OS 4_3_3 like Mac OS X; en-us) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8J2 Safari/6533.18.5",
    "Mozilla/5.0 (iPod; U; CPU iPhone OS 4_3_3 like Mac OS X; en-us) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8J2 Safari/6533.18.5",
    "Mozilla/5.0 (iPad; U; CPU OS 4_3_3 like Mac OS X; en-us) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8J2 Safari/6533.18.5",
    "Mozilla/5.0 (Linux; U; Android 2.3.7; en-us; Nexus One Build/FRF91) AppleWebKit/533.1 (KHTML, like Gecko) Version/4.0 Mobile Safari/533.1",
    "MQQBrowser/26 Mozilla/5.0 (Linux; U; Android 2.3.7; zh-cn; MB200 Build/GRJ22; CyanogenMod-7) AppleWebKit/533.1 (KHTML, like Gecko) Version/4.0 Mobile Safari/533.1",
    "Opera/9.80 (Android 2.3.4; Linux; Opera Mobi/build-1107180945; U; en-GB) Presto/2.8.149 Version/11.10",
    "Mozilla/5.0 (Linux; U; Android 3.0; en-us; Xoom Build/HRI39) AppleWebKit/534.13 (KHTML, like Gecko) Version/4.0 Safari/534.13",
    "Mozilla/5.0 (BlackBerry; U; BlackBerry 9800; en) AppleWebKit/534.1+ (KHTML, like Gecko) Version/6.0.0.337 Mobile Safari/534.1+",
    "Mozilla/5.0 (hp-tablet; Linux; hpwOS/3.0.0; U; en-US) AppleWebKit/534.6 (KHTML, like Gecko) wOSBrowser/233.70 Safari/534.6 TouchPad/1.0",
    "Mozilla/5.0 (SymbianOS/9.4; Series60/5.0 NokiaN97-1/20.0.019; Profile/MIDP-2.1 Configuration/CLDC-1.1) AppleWebKit/525 (KHTML, like Gecko) BrowserNG/7.1.18124",
    "Mozilla/5.0 (compatible; MSIE 9.0; Windows Phone OS 7.5; Trident/5.0; IEMobile/9.0; HTC; Titan)",
    "UCWEB7.0.2.37/28/999",
    "NOKIA5700/ UCWEB7.0.2.37/28/999",
    "Openwave/ UCWEB7.0.2.37/28/999",
    "Mozilla/4.0 (compatible; MSIE 6.0; ) Opera/UCWEB7.0.2.37/28/999",
    ]





url = 'https://offer.ebay.com/ws/eBayISAPI.dll?ViewBidsLogin&item={}&rt=nc&_trksid=p2047675.l2564'




class leastedData:
    def __init__(self):
        self.url = None
        self.itemid = None
        self.sale_time =self.price = self.headers = None
    def dataRequest(self):
        response = requests.get(url=self.url,headers = self.headers,proxies=proxies)
        # print(response.text)
        html = etree.HTML(response.text)
        for p in range(1,4):
            try:
                tr_List = html.xpath('//div[@class="BHbidSecBorderGrey"]/div/table[{}]//tr'.format(p))
                # print(tr_List)
                for td in tr_List[1:]:
                    self.sale_time = count_number =  self.price = "no data"
                    try:
                        self.sale_time = td.xpath('td[last()-1]/text()')[0]
                        count_number = td.xpath('td[last()-2]/text()')[0]
                        self.price = td.xpath('td[last()-3]/text()')[0]
                        cn.saleInfoInsert(itemid=self.itemid,price=self.price,Quantity=count_number,Date_of_Purchase=self.sale_time)
                    except Exception as y:
                        pass
                        # print(self.itemid,'内部读取错误',y)
            except Exception as e:
                pass
                # print('未获取到相关数据',e)
    def timeProcess(self):
            # 处理时间问题
            sales_date = self.sale_time.upper().repalce()
            month,days,year = sales_date.split("-")
            use_data = "20"+str(year) + "-" + time_dict(month) + "-" + str(days)
            print(use_data)
            standrd_sale_time = datetime.strptime(sales_date, '%Y-%m-%d')
            print(sales_date)
            now_time = datetime.now()#获取当前时间
            differ_time = (now_time - standrd_sale_time).days
    def priceProcess(self):
        # 处理价格
        price_number = re.findall(r'(\d+\.\d+)', self.price)
        if price_number:
            price_number = price_number[0]
            unit = re.findall(r'(\w+)', self.price)[0]

    def run(self,start,end):
        ws = load_workbook('de_sales.xlsx')
        wt = ws['Sheet1']
        for i in range(start,end):
            print(i)
            self.headers = {'User-Agent': random.choice(user_agent)}
            self.itemid = wt.cell(row=i,column=2).value
            self.url = url.format(self.itemid)
            self.dataRequest()

if __name__ == '__main__':
    leastedData().run(start=2,end=3549)
    # print(requests.get(url='http://httpbin.org/get').text)