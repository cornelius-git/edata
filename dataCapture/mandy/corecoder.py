
import  requests,random,pymysql
from lxml import etree
from openpyxl import load_workbook
url = "https://www.ebay.com/bin/purchaseHistory?item=143646288386&rt=nc&_trksid=p2047675.l2564"
proxy = '127.0.0.1:10808'
proxies = {
    'http': 'socks5://' + proxy,
    'https': 'socks5://' + proxy
}
db = pymysql.connect(host="127.0.0.1", user="controlarm",
                     password="123456", db="controlarm", port=3306,charset='utf8mb4')

# 使用cursor()方法获取操作游标
cur = db.cursor()
user_agent = [
    "Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_8; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50",
    "Mozilla/5.0 (Windows; U; Windows NT 6.1; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50",
    "Mozilla/5.0 (Windows NT 10.0; WOW64; rv:38.0) Gecko/20100101 Firefox/38.0",
    "Mozilla/5.0 (Windows NT 10.0; WOW64; Trident/7.0; .NET4.0C; .NET4.0E; .NET CLR 2.0.50727; .NET CLR 3.0.30729; .NET CLR 3.5.30729; InfoPath.3; rv:11.0) like Gecko",
    "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0)",
    "Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0)",
    "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1)",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.6; rv:2.0.1) Gecko/20100101 Firefox/4.0.1",
    "Mozilla/5.0 (Windows NT 6.1; rv:2.0.1) Gecko/20100101 Firefox/4.0.1",
    "Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; en) Presto/2.8.131 Version/11.11",
    "Opera/9.80 (Windows NT 6.1; U; en) Presto/2.8.131 Version/11.11",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_0) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Maxthon 2.0)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; TencentTraveler 4.0)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; The World)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; SE 2.X MetaSr 1.0; SE 2.X MetaSr 1.0; .NET CLR 2.0.50727; SE 2.X MetaSr 1.0)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; 360SE)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Avant Browser)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1)",
    "Mozilla/5.0 (iPhone; U; CPU iPhone OS 4_3_3 like Mac OS X; en-us) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8J2 Safari/6533.18.5",
    "Mozilla/5.0 (iPod; U; CPU iPhone OS 4_3_3 like Mac OS X; en-us) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8J2 Safari/6533.18.5",
    "Mozilla/5.0 (iPad; U; CPU OS 4_3_3 like Mac OS X; en-us) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8J2 Safari/6533.18.5",
    "Mozilla/5.0 (Linux; U; Android 2.3.7; en-us; Nexus One Build/FRF91) AppleWebKit/533.1 (KHTML, like Gecko) Version/4.0 Mobile Safari/533.1",
    "MQQBrowser/26 Mozilla/5.0 (Linux; U; Android 2.3.7; zh-cn; MB200 Build/GRJ22; CyanogenMod-7) AppleWebKit/533.1 (KHTML, like Gecko) Version/4.0 Mobile Safari/533.1",
    "Opera/9.80 (Android 2.3.4; Linux; Opera Mobi/build-1107180945; U; en-GB) Presto/2.8.149 Version/11.10",
    "Mozilla/5.0 (Linux; U; Android 3.0; en-us; Xoom Build/HRI39) AppleWebKit/534.13 (KHTML, like Gecko) Version/4.0 Safari/534.13",
    "Mozilla/5.0 (BlackBerry; U; BlackBerry 9800; en) AppleWebKit/534.1+ (KHTML, like Gecko) Version/6.0.0.337 Mobile Safari/534.1+",
    "Mozilla/5.0 (hp-tablet; Linux; hpwOS/3.0.0; U; en-US) AppleWebKit/534.6 (KHTML, like Gecko) wOSBrowser/233.70 Safari/534.6 TouchPad/1.0",
    "Mozilla/5.0 (SymbianOS/9.4; Series60/5.0 NokiaN97-1/20.0.019; Profile/MIDP-2.1 Configuration/CLDC-1.1) AppleWebKit/525 (KHTML, like Gecko) BrowserNG/7.1.18124",
    "Mozilla/5.0 (compatible; MSIE 9.0; Windows Phone OS 7.5; Trident/5.0; IEMobile/9.0; HTC; Titan)",
    "UCWEB7.0.2.37/28/999",
    "NOKIA5700/ UCWEB7.0.2.37/28/999",
    "Openwave/ UCWEB7.0.2.37/28/999",
    "Mozilla/4.0 (compatible; MSIE 6.0; ) Opera/UCWEB7.0.2.37/28/999",
    ]


class salerecoder:
    def __init__(self):
        pass
    def sqlreader(self,start,end):
        # 在数据库中读取数据
        wb = load_workbook("cq.xlsx")
        wt = wb["Sheet1"]
        for i in range(start,end):
            self.itemid = wt.cell(row=i,column=1).value
            print(self.itemid)
            sql1 = "SELECT * FROM coilversold WHERE itemid = {}".format(self.itemid)
            cur.execute(sql1)
            f = cur.fetchall()
            print(f)
            if len(f) > 0:
                # print("存在")
                continue
            else:
                print("抓取")
                self.recoder()
            self.recoder()

    def datawriter(self,price=None,quantity=None,deal_time=None,source=None):
        sql2 = "insert into coilversold(itemid,price,Quantity,Date_of_Purchase,source) values(%s,%s,%s,%s,%s)"
        cur.execute(sql2,(self.itemid,price,quantity,deal_time,source))
        db.commit()
    def recoder(self):
        headers = {
            'User-Agent': user_agent[random.randint(0, 35)]}
        rty = open("cq.text",mode='a+',encoding='utf-8')
        url = "https://www.ebay.com/bin/purchaseHistory?item={}&rt=nc&_trksid=p2047675.l2564".format(self.itemid)
        response = requests.get(url=url,proxies=proxies)
        # print(response.text)
        html = etree.HTML(response.text)

        recent_tr = html.xpath('//*[@id="mainContent"]//div[contains(@class,"fixed-price")]//tr')
        history_tr = html.xpath('//*[@id="mainContent"]//div[contains(@class,"app-table offer")][1]//tr')
        if recent_tr[1:]:
            source = "offer"
            rty.write(str(self.itemid) + "#" + source + "\n")
            for fg in recent_tr[1:]:
                deal_time = "".join(fg.xpath('td[last()]//text()'))
                quantity = "".join(fg.xpath('td[last()-1]//text()'))
                price = "".join(fg.xpath('td[last()-2]//text()'))
                print(price, quantity, deal_time)
                self.datawriter(price=price,quantity=quantity,deal_time=deal_time,source=source)
        else:
            self.datawriter(price="judge")
            rty.write(str(self.itemid)+"#"+"no data"+"\n")
        if history_tr[1:]:
            source = 'history'
            for fg in history_tr[1:]:
                deal_time = "".join(fg.xpath('td[last()]//text()'))
                quantity = "".join(fg.xpath('td[last()-1]//text()'))
                price = "".join(fg.xpath('td[last()-2]//text()'))
                print(price,quantity,deal_time)
                self.datawriter(price=price, quantity=quantity, deal_time=deal_time, source=source)
        else:
            pass

if __name__ == '__main__':
    salerecoder().sqlreader(start=6401,end=14498)